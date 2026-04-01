"""
펩리치 바이오펩톤 성분 분석 — 데이터 파싱 및 엑셀 생성
"""

from __future__ import annotations

import io
import re
from typing import Any, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constants import (
    AMINO_ACIDS,
    FREE_SUGARS,
    GENERAL_COMPONENTS,
    LAB_CATEGORY_MAP,
    MINERALS,
    NUCLEIC_ACIDS,
    SENSANG_ITEMS,
    SUMMARY_TEST_MAP,
    VITB_NAME_MAP,
    VITAMIN_B,
)

# ──────────────────────────────────────────────
# 스타일 상수
# ──────────────────────────────────────────────
FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
FONT_SECTION = Font(name="맑은 고딕", size=10, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=9, bold=True)
FONT_DATA = Font(name="맑은 고딕", size=9)
FONT_NOTE = Font(name="맑은 고딕", size=8, italic=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


# ──────────────────────────────────────────────
# 유틸리티
# ──────────────────────────────────────────────
def parse_result(value) -> Union[float, str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    try:
        return float(s)
    except ValueError:
        return s


def is_numeric(v) -> bool:
    return isinstance(v, (int, float))


def _set_cell(ws, row, col, value, font=None, alignment=None, border=None, fill=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    return cell


# ──────────────────────────────────────────────
# Phase 1: 파싱
# ──────────────────────────────────────────────
def classify_file(filename: str) -> str:
    if "의뢰품검사상세" in filename:
        return "lab"
    return "summary"


def parse_lab_files(file_bytes_list: list[tuple[str, bytes]]) -> list[dict]:
    """의뢰품검사상세 파일들을 파싱하여 레코드 리스트 반환."""
    records = []
    for _fname, fbytes in file_bytes_list:
        df = pd.read_excel(io.BytesIO(fbytes), engine="openpyxl")
        cols = list(df.columns)
        for _, row in df.iterrows():
            sample_name = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            category_raw = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ""
            is_header = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ""
            detail_item = str(row.iloc[10]).strip() if pd.notna(row.iloc[10]) else ""
            unit = str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else ""
            result_raw = row.iloc[12]

            if is_header == "Y":
                continue

            category = LAB_CATEGORY_MAP.get(category_raw, "unknown")

            # 일반성분이 검사항목으로 직접 나오는 경우 (상세검사항목이 비어 있음)
            # 예: 검사항목="TN (총질소)_DUMAS법", 상세검사항목=""
            if category == "general" and not detail_item:
                std = SUMMARY_TEST_MAP.get(category_raw)
                if std:
                    item_name = std
                else:
                    continue
            elif not detail_item:
                continue
            else:
                item_name = detail_item

            # 아미노산 항목명에서 (총)/(유리) 접미사 제거
            # 예: "Aspartic acid (총)" → "Aspartic acid"
            if category in ("taa", "faa"):
                item_name = re.sub(r"\s*\(총\)\s*$", "", item_name)
                item_name = re.sub(r"\s*\(유리\)\s*$", "", item_name)

            if category == "vitB":
                item_name = VITB_NAME_MAP.get(detail_item, item_name)
                if item_name == detail_item:
                    m = re.match(r"Vitamin B(\d+)", detail_item)
                    if m:
                        item_name = f"B{m.group(1)}"

            records.append({
                "sample_name": sample_name,
                "category": category,
                "item": item_name,
                "unit": unit,
                "result": parse_result(result_raw),
            })
    return records



def parse_summary_files(file_bytes_list: list[tuple[str, bytes]]) -> list[dict]:
    """엑셀정리파일들을 파싱하여 레코드 리스트 반환."""
    records = []
    for _fname, fbytes in file_bytes_list:
        df = pd.read_excel(io.BytesIO(fbytes), engine="openpyxl")
        sample_col = None
        test_col = None
        result_col = None
        unit_col = None

        for i, c in enumerate(df.columns):
            cs = str(c).strip()
            if "의뢰품명" in cs:
                sample_col = i
            elif "검사항목" in cs and "상세" not in cs:
                test_col = i
            elif "상세검사항목" in cs:
                pass
            elif "검사결과" in cs:
                result_col = i
            elif cs == "단위" or cs == "단위_1":
                unit_col = i

        if sample_col is None:
            for i, c in enumerate(df.columns):
                if "의뢰품" in str(c):
                    sample_col = i
                    break

        if test_col is None:
            for i, c in enumerate(df.columns):
                if "검사항목" in str(c):
                    test_col = i
                    break

        if result_col is None:
            for i, c in enumerate(df.columns):
                if "검사결과" in str(c) or "결과" in str(c):
                    result_col = i
                    break

        if sample_col is None or test_col is None or result_col is None:
            continue

        df.iloc[:, sample_col] = df.iloc[:, sample_col].ffill()

        for _, row in df.iterrows():
            sample_name = str(row.iloc[sample_col]).strip() if pd.notna(row.iloc[sample_col]) else ""
            test_name = str(row.iloc[test_col]).strip() if pd.notna(row.iloc[test_col]) else ""
            result_raw = row.iloc[result_col]
            unit = str(row.iloc[unit_col]).strip() if unit_col is not None and pd.notna(row.iloc[unit_col]) else ""

            if not sample_name or sample_name == "nan" or not test_name or test_name == "nan":
                continue

            std_name = SUMMARY_TEST_MAP.get(test_name, test_name)
            records.append({
                "sample_name": sample_name,
                "category": "general",
                "item": std_name,
                "unit": unit,
                "result": parse_result(result_raw),
            })
    return records


def prescan_files(file_bytes_list: list[tuple[str, bytes]]) -> dict:
    """파일들에서 시료명 추출."""
    lab_samples = set()
    summary_samples = set()
    for fname, fbytes in file_bytes_list:
        ftype = classify_file(fname)
        try:
            df = pd.read_excel(io.BytesIO(fbytes), engine="openpyxl")
            if ftype == "lab":
                for v in df.iloc[:, 0].dropna().unique():
                    lab_samples.add(str(v).strip())
            else:
                sample_col = None
                for i, c in enumerate(df.columns):
                    if "의뢰품" in str(c):
                        sample_col = i
                        break
                if sample_col is not None:
                    df.iloc[:, sample_col] = df.iloc[:, sample_col].ffill()
                    for v in df.iloc[:, sample_col].dropna().unique():
                        s = str(v).strip()
                        if s and s != "nan":
                            summary_samples.add(s)
        except Exception:
            continue
    return {
        "lab_samples": sorted(lab_samples),
        "summary_samples": sorted(summary_samples),
    }


def build_sample_data(
    lab_records: list[dict],
    summary_records: list[dict],
    sample_config: list[dict],
) -> dict[str, dict]:
    """시료별 데이터를 중첩 딕셔너리로 구축."""
    name_to_key = {}
    for sc in sample_config:
        key = sc["display_name"]
        for alias in sc.get("file_names", []):
            name_to_key[alias] = key
        name_to_key[key] = key

    data: dict[str, dict] = {}
    for sc in sample_config:
        key = sc["display_name"]
        data[key] = {
            "general": {},
            "free_sugar": {},
            "mineral": {},
            "nucleic_acid": {},
            "vitB": {},
            "faa": {},
            "taa": {},
            "organic_acid": {},
        }

    all_records = lab_records + summary_records
    for rec in all_records:
        key = name_to_key.get(rec["sample_name"])
        if key is None:
            continue
        cat = rec["category"]
        if cat not in data[key]:
            data[key][cat] = {}
        data[key][cat][rec["item"]] = rec["result"]

    return data


# ──────────────────────────────────────────────
# Phase 2: 엑셀 생성
# ──────────────────────────────────────────────
def generate_excel(
    sample_data: dict[str, dict],
    sensang_data: dict[str, dict],
    sample_config: list[dict],
    batch_date: str,
) -> bytes:
    wb = Workbook()
    samples = [sc["display_name"] for sc in sample_config]
    n = len(samples)

    _build_summary_sheet(wb, samples, sample_data, sensang_data, n)

    for sc in sample_config:
        name = sc["display_name"]
        sheet_name = sc.get("sheet_name", name)
        raw_name = sc.get("raw_material_name", "")
        _build_sample_sheet(wb, sheet_name, name, raw_name, sample_data, sensang_data)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_summary_sheet(wb, samples, sample_data, sensang_data, n):
    ws = wb.active
    ws.title = "성상, 성분"

    # 컬럼 너비
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 14
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12
    gap1 = 3 + n
    ws.column_dimensions[get_column_letter(gap1)].width = 2

    # === Row 1: 제목 ===
    _set_cell(ws, 1, 1, "바이오펩톤 성상 및 성분 분석", FONT_TITLE, ALIGN_LEFT)
    ws.row_dimensions[1].height = 22

    # === 성상 섹션 (Row 3~10) ===
    _set_cell(ws, 3, 1, "성상", FONT_SECTION, ALIGN_CENTER, THIN_BORDER, FILL_SECTION)
    for i, s in enumerate(samples):
        _set_cell(ws, 3, 3 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)

    _set_cell(ws, 4, 2, "사진", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
    ws.row_dimensions[4].height = 70

    sensang_items_keys = ["pH", "탁도", "색도 (L)", "색도 (a)", "색도 (b)"]
    sensang_json_keys = ["pH", "탁도", "색도_L", "색도_a", "색도_b"]
    for ri, (label, jkey) in enumerate(zip(sensang_items_keys, sensang_json_keys)):
        row = 5 + ri
        _set_cell(ws, row, 2, label, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            sd = sensang_data.get(s, {})
            val = sd.get(jkey, "")
            if val != "":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            _set_cell(ws, row, 3 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    _set_cell(ws, 10, 2, "* pH, 탁도, 색도는 2% solution에서 측정", FONT_NOTE, ALIGN_LEFT)

    # === 성분 분석: 1. 일반성분 + 2. 유리당 (Row 12~) ===
    sec2_start = gap1 + 1  # 유리당 시작 컬럼

    _set_cell(ws, 12, 1, "성분 분석", FONT_SECTION, ALIGN_CENTER, THIN_BORDER, FILL_SECTION)
    _set_cell(ws, 12, 2, "1. 일반 성분 (%)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)
    _set_cell(ws, 12, sec2_start, "2. 유리당 (mg/kg)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)

    # 헤더 행 13
    for i, s in enumerate(samples):
        _set_cell(ws, 13, 3 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        _set_cell(ws, 13, sec2_start + 1 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)

    # 일반성분 데이터 (Row 14~21)
    for ri, comp in enumerate(GENERAL_COMPONENTS):
        row = 14 + ri
        _set_cell(ws, row, 2, comp, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            val = sample_data.get(s, {}).get("general", {}).get(comp, "")
            _set_cell(ws, row, 3 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # 유리당 데이터 (Row 14~18)
    for ri, sugar in enumerate(FREE_SUGARS):
        row = 14 + ri
        _set_cell(ws, row, sec2_start, sugar, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            val = sample_data.get(s, {}).get("free_sugar", {}).get(sugar, "")
            _set_cell(ws, row, sec2_start + 1 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # === 3. 미네랄 + 4. 핵산 + 5. 비타민 B (Row 24~) ===
    row_sec3 = 23
    sec3_label_col = 2
    sec4_start = gap1 + 1
    gap2 = sec4_start + n + 1
    sec5_start = gap2 + 1

    ws.column_dimensions[get_column_letter(gap1)].width = 2
    ws.column_dimensions[get_column_letter(gap2)].width = 2 if gap2 <= 26 else 2

    _set_cell(ws, row_sec3, 2, "3. 미네랄 (mg/kg)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)
    _set_cell(ws, row_sec3, sec4_start, "4. 핵산 (%)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)
    _set_cell(ws, row_sec3, sec5_start, "5. 비타민 B (mg/kg)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)

    # 헤더
    for i, s in enumerate(samples):
        _set_cell(ws, row_sec3 + 1, 3 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        _set_cell(ws, row_sec3 + 1, sec4_start + 1 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        _set_cell(ws, row_sec3 + 1, sec5_start + 1 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)

    # 미네랄
    for ri, item in enumerate(MINERALS):
        row = row_sec3 + 2 + ri
        _set_cell(ws, row, 2, item, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            val = sample_data.get(s, {}).get("mineral", {}).get(item, "")
            _set_cell(ws, row, 3 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # 핵산
    for ri, item in enumerate(NUCLEIC_ACIDS):
        row = row_sec3 + 2 + ri
        _set_cell(ws, row, sec4_start, item, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            val = sample_data.get(s, {}).get("nucleic_acid", {}).get(item, "")
            _set_cell(ws, row, sec4_start + 1 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # 비타민 B
    for ri, item in enumerate(VITAMIN_B):
        row = row_sec3 + 2 + ri
        _set_cell(ws, row, sec5_start, item, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            val = sample_data.get(s, {}).get("vitB", {}).get(item, "")
            _set_cell(ws, row, sec5_start + 1 + i, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # === 6. 아미노산 (Row 34~) ===
    aa_start_row = row_sec3 + 2 + max(len(MINERALS), len(NUCLEIC_ACIDS), len(VITAMIN_B)) + 2

    _set_cell(ws, aa_start_row, 2, "6. 총/유리 아미노산 (%)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)

    # 4블록: 유리AA함량, 유리AA조성비, 총AA함량, 총AA조성비
    blk_labels = ["■ 유리아미노산 함량", "■ 유리아미노산 조성비", "■ 총아미노산 함량", "■ 총아미노산 조성비"]
    blk_cats = ["faa", "faa_ratio", "taa", "taa_ratio"]
    blk_starts = []
    col_cursor = 2
    for bi in range(4):
        blk_starts.append(col_cursor)
        _set_cell(ws, aa_start_row + 1, col_cursor, blk_labels[bi], FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_SECTION)
        # sub-header: "(%) " + sample names
        _set_cell(ws, aa_start_row + 2, col_cursor, "(%)", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            _set_cell(ws, aa_start_row + 2, col_cursor + 1 + i, s, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        col_cursor += 1 + n + 1  # label + n samples + gap

    # 아미노산 데이터 행
    data_start = aa_start_row + 3
    for ai, aa in enumerate(AMINO_ACIDS):
        row = data_start + ai
        for bi, (bstart, bcat) in enumerate(zip(blk_starts, blk_cats)):
            _set_cell(ws, row, bstart, aa, FONT_DATA, ALIGN_LEFT, THIN_BORDER)
            for i, s in enumerate(samples):
                col = bstart + 1 + i
                if bcat == "faa":
                    val = sample_data.get(s, {}).get("faa", {}).get(aa, "")
                    _set_cell(ws, row, col, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
                elif bcat == "taa":
                    val = sample_data.get(s, {}).get("taa", {}).get(aa, "")
                    _set_cell(ws, row, col, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
                elif bcat == "faa_ratio":
                    src_col = blk_starts[0] + 1 + i
                    sum_row = data_start + len(AMINO_ACIDS)
                    src_letter = get_column_letter(src_col)
                    sum_letter = get_column_letter(src_col)
                    val_cell = sample_data.get(s, {}).get("faa", {}).get(aa, "")
                    if is_numeric(val_cell) and val_cell != 0:
                        formula = f"={src_letter}{row}/${src_letter}${sum_row}*100"
                        _set_cell(ws, row, col, formula, FONT_DATA, ALIGN_CENTER, THIN_BORDER, number_format="0.00")
                    else:
                        _set_cell(ws, row, col, 0 if is_numeric(val_cell) else val_cell, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
                elif bcat == "taa_ratio":
                    src_col = blk_starts[2] + 1 + i
                    sum_row = data_start + len(AMINO_ACIDS)
                    src_letter = get_column_letter(src_col)
                    val_cell = sample_data.get(s, {}).get("taa", {}).get(aa, "")
                    if is_numeric(val_cell) and val_cell != 0:
                        formula = f"={src_letter}{row}/${src_letter}${sum_row}*100"
                        _set_cell(ws, row, col, formula, FONT_DATA, ALIGN_CENTER, THIN_BORDER, number_format="0.00")
                    else:
                        _set_cell(ws, row, col, 0 if is_numeric(val_cell) else val_cell, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # 합계 행
    sum_row = data_start + len(AMINO_ACIDS)
    for bi, (bstart, bcat) in enumerate(zip(blk_starts, blk_cats)):
        _set_cell(ws, sum_row, bstart, "합계", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        for i, s in enumerate(samples):
            col = bstart + 1 + i
            cl = get_column_letter(col)
            formula = f"=SUM({cl}{data_start}:{cl}{sum_row - 1})"
            _set_cell(ws, sum_row, col, formula, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER, number_format="0.00")


def _build_sample_sheet(wb, sheet_name, display_name, raw_name, sample_data, sensang_data):
    ws = wb.create_sheet(title=sheet_name)
    has_raw = bool(raw_name)
    data_cols = 2 if has_raw else 1  # 원료 + BIO 또는 BIO만

    # 컬럼 너비
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    if has_raw:
        ws.column_dimensions["D"].width = 12

    # === Row 1: 제목 ===
    _set_cell(ws, 1, 1, f"{display_name} 성상 및 성분 분석", FONT_TITLE, ALIGN_LEFT)

    # === 성상 (Row 3~9) ===
    _set_cell(ws, 3, 1, "성상", FONT_SECTION, ALIGN_CENTER, THIN_BORDER, FILL_SECTION)
    if has_raw:
        _set_cell(ws, 3, 3, raw_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        _set_cell(ws, 3, 4, display_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
    else:
        _set_cell(ws, 3, 3, display_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)

    sensang_items_keys = ["pH", "탁도", "색도 (L)", "색도 (a)", "색도 (b)"]
    sensang_json_keys = ["pH", "탁도", "색도_L", "색도_a", "색도_b"]
    for ri, (label, jkey) in enumerate(zip(sensang_items_keys, sensang_json_keys)):
        row = 4 + ri
        _set_cell(ws, row, 2, label, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        sd = sensang_data.get(display_name, {})
        bio_val = sd.get(jkey, "")
        if bio_val != "":
            try:
                bio_val = float(bio_val)
            except (ValueError, TypeError):
                pass
        if has_raw:
            _set_cell(ws, row, 3, "", FONT_DATA, ALIGN_CENTER, THIN_BORDER)
            _set_cell(ws, row, 4, bio_val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
        else:
            _set_cell(ws, row, 3, bio_val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    _set_cell(ws, 9, 2, "* pH, 탁도, 색도는 2% solution에서 측정", FONT_NOTE, ALIGN_LEFT)

    # === 성분 분석 (Row 11~) ===
    r = 11
    bio_col = 4 if has_raw else 3
    raw_col = 3 if has_raw else None

    # -- 6개 섹션을 가로로 3개씩 2줄로 배치 --
    sections_row1 = [
        ("1. 일반 성분 (%)", GENERAL_COMPONENTS, "general"),
        ("2. 유리당 (mg/kg)", FREE_SUGARS, "free_sugar"),
        ("3. 미네랄 (mg/kg)", MINERALS, "mineral"),
    ]
    sections_row2 = [
        ("4. 핵산 (%)", NUCLEIC_ACIDS, "nucleic_acid"),
        ("5. 비타민 B (mg/kg)", VITAMIN_B, "vitB"),
    ]

    def _write_section_block(ws, start_row, start_col, title, items, cat_key, display_name, raw_name, sample_data, has_raw):
        _set_cell(ws, start_row, start_col, title, FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)
        # 헤더
        hr = start_row + 1
        if has_raw:
            _set_cell(ws, hr, start_col + 1, raw_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
            _set_cell(ws, hr, start_col + 2, display_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        else:
            _set_cell(ws, hr, start_col + 1, display_name, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)

        for ri, item in enumerate(items):
            row = hr + 1 + ri
            _set_cell(ws, row, start_col, item, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
            bio_val = sample_data.get(display_name, {}).get(cat_key, {}).get(item, "")
            if has_raw:
                _set_cell(ws, row, start_col + 1, "", FONT_DATA, ALIGN_CENTER, THIN_BORDER)
                _set_cell(ws, row, start_col + 2, bio_val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
            else:
                _set_cell(ws, row, start_col + 1, bio_val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)

        return hr + 1 + len(items)

    _set_cell(ws, r, 1, "성분 분석", FONT_SECTION, ALIGN_CENTER, THIN_BORDER, FILL_SECTION)

    col_gap = data_cols + 2  # label + data_cols + gap
    col = 2
    max_end = r
    for title, items, cat in sections_row1:
        ws.column_dimensions[get_column_letter(col)].width = 14
        for dc in range(1, data_cols + 1):
            ws.column_dimensions[get_column_letter(col + dc)].width = 12
        end = _write_section_block(ws, r, col, title, items, cat, display_name, raw_name, sample_data, has_raw)
        max_end = max(max_end, end)
        col += col_gap

    r2 = max_end + 1
    col = 2
    for title, items, cat in sections_row2:
        ws.column_dimensions[get_column_letter(col)].width = 14
        for dc in range(1, data_cols + 1):
            ws.column_dimensions[get_column_letter(col + dc)].width = 12
        end = _write_section_block(ws, r2, col, title, items, cat, display_name, raw_name, sample_data, has_raw)
        col += col_gap

    # === 아미노산 섹션 ===
    aa_row = r2 + max(len(NUCLEIC_ACIDS), len(VITAMIN_B)) + 3
    _set_cell(ws, aa_row, 2, "6. 총/유리 아미노산 (%)", FONT_SECTION, ALIGN_LEFT, THIN_BORDER, FILL_SECTION)

    # 4블록: 함량(유리/총), 조성비(유리/총)
    blk_labels = ["함량(%)", None, "조성비(%)", None]
    blk_sub = ["유리 아미노산", "총 아미노산", "유리 아미노산", "총 아미노산"]
    blk_cats = ["faa", "taa", "faa_ratio", "taa_ratio"]

    # 헤더 구성
    hdr_row1 = aa_row + 1
    hdr_row2 = aa_row + 2

    col_cursor = 3
    blk_col_starts = []
    for bi in range(4):
        blk_col_starts.append(col_cursor)
        if bi == 0:
            _set_cell(ws, hdr_row1, col_cursor, "함량(%)", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        elif bi == 2:
            _set_cell(ws, hdr_row1, col_cursor, "조성비(%)", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        _set_cell(ws, hdr_row2, col_cursor, blk_sub[bi], FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
        ws.column_dimensions[get_column_letter(col_cursor)].width = 14
        col_cursor += 1

    data_start = hdr_row2 + 1
    for ai, aa in enumerate(AMINO_ACIDS):
        row = data_start + ai
        _set_cell(ws, row, 2, aa, FONT_DATA, ALIGN_LEFT, THIN_BORDER)
        for bi, (bcol, bcat) in enumerate(zip(blk_col_starts, blk_cats)):
            if bcat == "faa":
                val = sample_data.get(display_name, {}).get("faa", {}).get(aa, "")
                _set_cell(ws, row, bcol, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
            elif bcat == "taa":
                val = sample_data.get(display_name, {}).get("taa", {}).get(aa, "")
                _set_cell(ws, row, bcol, val, FONT_DATA, ALIGN_CENTER, THIN_BORDER)
            elif bcat == "faa_ratio":
                src_col = blk_col_starts[0]
                sum_row = data_start + len(AMINO_ACIDS)
                cl = get_column_letter(src_col)
                val_cell = sample_data.get(display_name, {}).get("faa", {}).get(aa, "")
                if is_numeric(val_cell) and val_cell != 0:
                    formula = f"={cl}{row}/${cl}${sum_row}*100"
                    _set_cell(ws, row, bcol, formula, FONT_DATA, ALIGN_CENTER, THIN_BORDER, number_format="0.00")
                else:
                    _set_cell(ws, row, bcol, 0 if is_numeric(val_cell) else "", FONT_DATA, ALIGN_CENTER, THIN_BORDER)
            elif bcat == "taa_ratio":
                src_col = blk_col_starts[1]
                sum_row = data_start + len(AMINO_ACIDS)
                cl = get_column_letter(src_col)
                val_cell = sample_data.get(display_name, {}).get("taa", {}).get(aa, "")
                if is_numeric(val_cell) and val_cell != 0:
                    formula = f"={cl}{row}/${cl}${sum_row}*100"
                    _set_cell(ws, row, bcol, formula, FONT_DATA, ALIGN_CENTER, THIN_BORDER, number_format="0.00")
                else:
                    _set_cell(ws, row, bcol, 0 if is_numeric(val_cell) else "", FONT_DATA, ALIGN_CENTER, THIN_BORDER)

    # 합계 행
    sum_row = data_start + len(AMINO_ACIDS)
    _set_cell(ws, sum_row, 2, "합계", FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER)
    for bi, bcol in enumerate(blk_col_starts):
        cl = get_column_letter(bcol)
        formula = f"=SUM({cl}{data_start}:{cl}{sum_row - 1})"
        _set_cell(ws, sum_row, bcol, formula, FONT_HEADER, ALIGN_CENTER, THIN_BORDER, FILL_HEADER, number_format="0.00")


# ──────────────────────────────────────────────
# 메인 진입점
# ──────────────────────────────────────────────
def process_all(
    file_bytes_list: list[tuple[str, bytes]],
    sample_config: list[dict],
    sensang_data: dict[str, dict],
    batch_date: str,
) -> tuple[bytes, dict]:
    lab_files = [(f, b) for f, b in file_bytes_list if classify_file(f) == "lab"]
    summary_files = [(f, b) for f, b in file_bytes_list if classify_file(f) == "summary"]

    lab_records = parse_lab_files(lab_files)
    summary_records = parse_summary_files(summary_files)
    sample_data = build_sample_data(lab_records, summary_records, sample_config)

    excel_bytes = generate_excel(sample_data, sensang_data, sample_config, batch_date)

    summary_info = {
        "sample_count": len(sample_config),
        "lab_records": len(lab_records),
        "summary_records": len(summary_records),
        "samples": [sc["display_name"] for sc in sample_config],
    }
    return excel_bytes, summary_info
